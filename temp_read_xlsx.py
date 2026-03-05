import openpyxl
p = rc:\Users\abdia\Downloads\Справочник полей HSE (1).xlsx
wb = openpyxl.load_workbook(p, data_only=True)
print('SHEETS:', wb.sheetnames)
for s in wb.sheetnames:
 ws = wb[s]
 print('\n==', s, '==')
 max_rows = min(ws.max_row, 40)
 max_cols = min(ws.max_column, 10)
 for r in range(1, max_rows + 1):
 vals = [ws.cell(r, c).value for c in range(1, max_cols + 1)]
 if any(v is not None and str(v).strip() != '' for v in vals):
 print('\t'.join('' if v is None else str(v) for v in vals))
